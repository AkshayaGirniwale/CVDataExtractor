import os
from flask import Flask, render_template, request, redirect, flash, send_file
from pdfminer.high_level import extract_text
from docx import Document
import re
import io
from openpyxl import load_workbook, Workbook
import shutil
import tempfile

app = Flask(__name__)

UPLOAD_FOLDER = 'upload'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
CV_DATA_FILE = 'cv_data.xlsx'

def extract_text_from_docx(docx_file):
    doc = Document(docx_file)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def extract_data_from_cv(cv_file):
    # Extract text from PDF or Word document
    if cv_file.filename.endswith('.pdf'):
        text = extract_text(io.BytesIO(cv_file.read()))
    elif cv_file.filename.endswith('.docx'):
        text = extract_text_from_docx(io.BytesIO(cv_file.read()))
    else:
        return None, None, None
    
    # Extract email and phone number using regex
    email = re.search(r'[\w\.-]+@[\w\.-]+', text).group(0)
    phone = re.search(r'[\+\(]?[1-9][0-9 .\-\(\)]{8,}[0-9]', text).group(0)
    
    return email, phone, text

def create_excel(email, phone, text, output_file):
    # Check if the output file already exists
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Email', 'Phone', 'Text'])

    # Sanitize text to remove unsupported characters
    sanitized_text = sanitize_text(text)

    # Append new data to the worksheet
    ws.append([email, phone, sanitized_text])

    # Save the workbook
    wb.save(output_file)
    return output_file

def sanitize_text(text):
    # Define a list of unsupported characters
    unsupported_chars = ['\x0c']  # Add other unsupported characters if needed
    
    # Remove unsupported characters from the text
    sanitized_text = ''.join(char for char in text if char not in unsupported_chars)
    
    return sanitized_text

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'cv_file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        
        cv_file = request.files['cv_file']
        if cv_file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        
        # Check if cv_data.xlsx exists in the upload folder
        cv_data_path = os.path.join(app.config['UPLOAD_FOLDER'], CV_DATA_FILE)
        if os.path.exists(cv_data_path):
            output_file = cv_data_path
            append_data = True
        else:
            output_file = os.path.join(app.config['UPLOAD_FOLDER'], CV_DATA_FILE)
            append_data = False
        
        email, phone, text = extract_data_from_cv(cv_file)
        if email and phone and text:
            create_excel(email, phone, text, output_file)
        
        if append_data:
            return redirect(request.url)
        else:
            # Serve the file for download
            return send_file(output_file, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
